import os
from openpyxl import load_workbook
from openpyxl import Workbook
import re
from pathlib import Path
import gc

# ==================== 配置参数 ====================
MAX_DISPLAY_COLS = 0  # 0表示显示全部非空白列，设置具体数字则限制显示前N列
MAX_DISPLAY_ROWS = 20  # 控制台最多显示的行数
MAX_CELL_WIDTH = 30    # 每个单元格最大显示宽度（字符数）
CONSECUTIVE_EMPTY_THRESHOLD = 100  # 连续空白行阈值
DEFAULT_MATCH_MODE = "exact"  # 默认匹配模式: "fuzzy"(模糊) 或 "exact"(精确)
HIDE_EMPTY_COLUMNS = True  # 是否自动隐藏完全空白的列
# =================================================

# 全局变量
current_match_mode = DEFAULT_MATCH_MODE  # 当前匹配模式

def get_folder_path():
    """获取文件夹路径，支持重试"""
    while True:
        print("\n提示：可以直接将文件夹拖拽到窗口，然后按回车")
        print("      或输入 'q' 退出程序")
        folder_path = input("\n请输入文件夹路径：").strip()
        
        if folder_path.lower() == 'q':
            return None
        
        folder_path = folder_path.strip('"').strip("'").rstrip()
        
        if os.path.isdir(folder_path):
            return folder_path
        else:
            print(f"❌ 错误：文件夹不存在")
            print(f"   您输入的路径是：{folder_path}")
            continue

def get_excel_files(folder_path):
    """获取文件夹下所有Excel文件"""
    excel_extensions = {'.xlsx', '.xls'}
    excel_files = []
    
    for file in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file)
        if os.path.isfile(file_path):
            ext = os.path.splitext(file)[1].lower()
            if ext in excel_extensions:
                excel_files.append(file)
    
    return sorted(excel_files)

def format_file_size(size_bytes):
    """格式化文件大小显示"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.1f}{unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f}TB"

def select_excel_file(excel_files, folder_path):
    """让用户选择要读取的Excel文件"""
    if not excel_files:
        print("\n❌ 该文件夹下没有找到 Excel 文件（.xlsx 或 .xls）")
        return None
    
    while True:
        print(f"\n{'='*70}")
        print(f"找到 {len(excel_files)} 个 Excel 文件：")
        print(f"{'='*70}")
        
        for idx, filename in enumerate(excel_files, start=1):
            file_path = os.path.join(folder_path, filename)
            file_size = os.path.getsize(file_path)
            size_str = format_file_size(file_size)
            print(f"  {idx}. {filename} ({size_str})")
        
        print(f"  0. 重新选择文件夹")
        print(f"  q. 退出程序")
        
        choice = input(f"\n请选择要读取的文件（输入序号 1-{len(excel_files)}）：").strip()
        
        if choice.lower() == 'q':
            return None
        elif choice == '0':
            return 'RESELECT_FOLDER'
        elif choice.isdigit():
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(excel_files):
                return excel_files[choice_idx]
            else:
                print(f"❌ 序号超出范围，请输入 1 ~ {len(excel_files)}")
        else:
            print("❌ 输入无效，请输入序号")

def show_loading_progress(current, total, prefix='', suffix='', bar_length=40):
    """显示进度条"""
    if total == 0:
        return
    
    percent = current / total
    arrow = '█' * int(round(percent * bar_length))
    spaces = ' ' * (bar_length - len(arrow))
    percent_display = percent * 100
    
    if percent_display < 30:
        color = "🟡"
    elif percent_display < 80:
        color = "🔵"
    else:
        color = "🟢"
    
    print(f"\r{color} {prefix} |{arrow}{spaces}| {percent_display:.1f}% {suffix}", end='', flush=True)
    
    if current == total:
        print()

def is_row_empty(row):
    """判断一行是否为空（所有单元格都是None或空字符串）"""
    if not row:
        return True
    for cell in row:
        if cell is not None and str(cell).strip() != '':
            return False
    return True

def is_column_empty_in_rows(rows, col_index):
    """判断某一列在指定的行中是否全部为空"""
    for row in rows:
        if col_index < len(row) and row[col_index] is not None and str(row[col_index]).strip() != '':
            return False
    return True

def filter_empty_columns(headers, matched_rows):
    """过滤掉完全空白的列"""
    if not matched_rows or not headers:
        return headers, [], matched_rows
    
    # 获取所有匹配的行数据
    rows_data = [row for _, row in matched_rows]
    
    # 找出哪些列有数据
    non_empty_cols = []
    for col_idx in range(len(headers)):
        if not is_column_empty_in_rows(rows_data, col_idx):
            non_empty_cols.append(col_idx)
    
    # 如果没有非空列，返回原数据
    if len(non_empty_cols) == 0:
        return headers, list(range(len(headers))), matched_rows
    
    # 过滤表头
    filtered_headers = [headers[i] for i in non_empty_cols]
    
    # 过滤数据行
    filtered_rows = []
    for row_num, row in matched_rows:
        filtered_row = [row[i] if i < len(row) else None for i in non_empty_cols]
        filtered_rows.append((row_num, filtered_row))
    
    # 返回过滤后的结果和原始列索引映射
    return filtered_headers, non_empty_cols, filtered_rows

def match_cell_value(cell_value, search_term, match_mode):
    """根据匹配模式判断单元格是否匹配"""
    if cell_value is None:
        return False
    
    cell_str = str(cell_value)
    search_lower = search_term.lower()
    
    if match_mode == "fuzzy":
        # 模糊匹配：包含即可
        return search_lower in cell_str.lower()
    else:  # exact
        # 精确匹配：完全相等（忽略前后空格）
        return cell_str.strip().lower() == search_lower

def get_sheet_info_smart(file_path, sheet_name):
    """智能获取sheet的实际数据行数（提前终止扫描）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        total_rows = 0
        empty_rows = 0
        consecutive_empty_rows = 0
        first_row = True
        headers = None
        stop_early = False
        
        for row in ws.iter_rows(values_only=True):
            if first_row:
                headers = row
                first_row = False
                continue
            
            if not is_row_empty(row):
                total_rows += 1
                consecutive_empty_rows = 0
            else:
                empty_rows += 1
                consecutive_empty_rows += 1
            
            if consecutive_empty_rows >= CONSECUTIVE_EMPTY_THRESHOLD and total_rows > 0:
                print(f"   ⚡ 检测到连续 {CONSECUTIVE_EMPTY_THRESHOLD} 行空白，提前终止扫描")
                print(f"   ✅ 共扫描 {total_rows + empty_rows + 1:,} 行，其中有效数据 {total_rows:,} 行")
                stop_early = True
                break
            
            scanned_rows = total_rows + empty_rows
            if scanned_rows > 0 and scanned_rows % 50000 == 0:
                print(f"   已扫描 {scanned_rows:,} 行，有效数据 {total_rows:,} 行...")
        
        if not stop_early:
            print(f"   ✅ 扫描完成，共扫描 {total_rows + empty_rows:,} 行，有效数据 {total_rows:,} 行")
        
        wb.close()
        return total_rows, empty_rows, len(headers) if headers else 0, stop_early
        
    except Exception as e:
        print(f"   统计失败：{e}")
        return 0, 0, 0, False

def select_sheet_with_smart_filter(file_path, sheet_names):
    """优化的sheet选择，智能过滤空白行"""
    print(f"\n正在智能分析Sheet信息（检测到连续空白行会自动停止）...")
    
    sheet_info = []
    for idx, name in enumerate(sheet_names, start=1):
        print(f"\n  📊 分析 {idx}/{len(sheet_names)}: {name}...")
        valid_rows, empty_rows, col_count, early_stop = get_sheet_info_smart(file_path, name)
        sheet_info.append((name, valid_rows, empty_rows, col_count, early_stop))
        gc.collect()
    
    while True:
        print(f"\n{'='*80}")
        print(f"可用的 Sheet 页（共 {len(sheet_names)} 个）：")
        print(f"{'='*80}")
        
        for idx, (name, valid_rows, empty_rows, col_count, early_stop) in enumerate(sheet_info, start=1):
            if valid_rows > 0:
                info = f"有效数据:{valid_rows:,}行"
                if empty_rows > 0:
                    info += f" (跳过{empty_rows:,}个空白行)"
                if early_stop:
                    info += " ⚡智能截断"
                info += f", 列数:{col_count}"
                
                if valid_rows > 100000:
                    info += " ⚠️大文件"
                print(f"  {idx}. {name} ({info})")
            else:
                print(f"  {idx}. {name} (⚠️无有效数据，全是空白行)")
        
        choice = input("\n请选择 Sheet 页（输入序号/完整名称，或输入 'q' 返回）：").strip()
        
        if choice.lower() == 'q':
            return None
        
        if choice.isdigit():
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(sheet_names):
                return sheet_names[choice_idx]
            else:
                print(f"❌ 序号超出范围，请输入 1 ~ {len(sheet_names)}")
        else:
            if choice in sheet_names:
                return choice
            else:
                print(f"❌ 未找到名为 '{choice}' 的 Sheet")

def search_in_sheet_smart(file_path, sheet_name, search_term, match_mode):
    """智能搜索（提前终止扫描空白行）"""
    mode_text = "模糊匹配" if match_mode == "fuzzy" else "精确匹配"
    print(f"\n🔍 正在搜索 '{search_term}'（{mode_text}，智能跳过空白行）...")
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        # 第一遍：智能统计有效行数
        print("   正在统计有效数据行数...")
        valid_rows_count = 0
        consecutive_empty_rows = 0
        first_row = True
        headers = None
        
        for row in ws.iter_rows(values_only=True):
            if first_row:
                headers = row
                first_row = False
                continue
            
            if not is_row_empty(row):
                valid_rows_count += 1
                consecutive_empty_rows = 0
            else:
                consecutive_empty_rows += 1
            
            if consecutive_empty_rows >= CONSECUTIVE_EMPTY_THRESHOLD and valid_rows_count > 0:
                print(f"   ⚡ 检测到连续 {CONSECUTIVE_EMPTY_THRESHOLD} 行空白，提前终止统计")
                break
        
        print(f"   共发现 {valid_rows_count:,} 行有效数据")
        
        # 第二遍：智能搜索
        wb.close()
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        matched_rows = []
        current_row = 0
        valid_row_index = 0
        consecutive_empty_rows = 0
        first_row = True
        stop_early = False
        
        for row in ws.iter_rows(values_only=True):
            current_row += 1
            
            if first_row:
                first_row = False
                continue
            
            if is_row_empty(row):
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= CONSECUTIVE_EMPTY_THRESHOLD and valid_row_index > 0:
                    print(f"   ⚡ 检测到连续 {CONSECUTIVE_EMPTY_THRESHOLD} 行空白，提前终止搜索")
                    stop_early = True
                    break
                continue
            
            consecutive_empty_rows = 0
            valid_row_index += 1
            
            if valid_row_index % 5000 == 0 or valid_row_index == valid_rows_count:
                show_loading_progress(valid_row_index, valid_rows_count, 
                                     prefix='搜索进度', 
                                     suffix=f'({valid_row_index:,}/{valid_rows_count:,} 行)')
            
            if row:
                for cell in row:
                    if match_cell_value(cell, search_term, match_mode):
                        matched_rows.append((current_row, row))
                        break
        
        if stop_early:
            print(f"   ✅ 提前终止，实际搜索 {valid_row_index:,} 行有效数据")
        else:
            print()
        
        wb.close()
        return headers, matched_rows, valid_rows_count
        
    except Exception as e:
        print(f"\n❌ 搜索失败：{e}")
        import traceback
        traceback.print_exc()
        return None, [], 0

def display_results_with_config(headers, matched_rows, search_term, total_valid_rows, match_mode):
    """根据配置显示搜索结果（自动隐藏空白列）"""
    global MAX_DISPLAY_COLS, MAX_DISPLAY_ROWS, MAX_CELL_WIDTH, HIDE_EMPTY_COLUMNS
    
    if not matched_rows:
        mode_text = "模糊" if match_mode == "fuzzy" else "精确"
        print(f"\n❌ 在 {total_valid_rows:,} 行有效数据中，未找到{mode_text}匹配 '{search_term}' 的内容。")
        return False
    
    # 过滤空白列
    original_col_count = len(headers) if headers else 0
    if HIDE_EMPTY_COLUMNS and headers:
        filtered_headers, col_mapping, filtered_rows = filter_empty_columns(headers, matched_rows)
        # 更新显示数据
        display_headers = filtered_headers
        display_rows = filtered_rows
        hidden_count = original_col_count - len(display_headers)
        
        if hidden_count > 0:
            print(f"\n✨ 自动隐藏了 {hidden_count} 个完全空白的列，当前显示 {len(display_headers)} 个有数据的列")
    else:
        display_headers = headers
        display_rows = matched_rows
        hidden_count = 0
    
    mode_text = "模糊" if match_mode == "fuzzy" else "精确"
    print(f"\n{'='*80}")
    print(f"✅ 在 {total_valid_rows:,} 行有效数据中，找到 {len(matched_rows):,} 条{mode_text}匹配结果：")
    print(f"{'='*80}\n")
    
    # 准备表头
    if display_headers:
        header_display = [str(h) if h is not None else f"列{i+1}" for i, h in enumerate(display_headers)]
    else:
        header_display = [f"列{i+1}" for i in range(len(display_rows[0][1]) if display_rows else 0)]
    
    total_cols = len(header_display)
    
    # 确定显示哪些列
    if MAX_DISPLAY_COLS == 0:
        # 显示全部非空白列
        display_cols = total_cols
        if total_cols > 0:
            print(f"📊 显示全部 {total_cols} 列（已自动隐藏空白列）")
    else:
        display_cols = min(MAX_DISPLAY_COLS, total_cols)
        if display_cols < total_cols:
            print(f"📊 显示前 {display_cols} 列（共 {total_cols} 列有数据）")
            print(f"   💡 提示：如需显示全部列，请输入 '/showall'")
    
    if display_cols == 0:
        print("⚠️ 没有可显示的列（所有列都是空白的）")
        return True
    
    header_display = header_display[:display_cols]
    
    # 计算列宽
    col_widths = []
    for i, h in enumerate(header_display):
        width = min(len(h), MAX_CELL_WIDTH)
        for _, row in display_rows[:MAX_DISPLAY_ROWS]:
            if i < len(row) and row[i] is not None:
                cell_str = str(row[i])
                width = max(width, min(len(cell_str), MAX_CELL_WIDTH))
        col_widths.append(width)
    
    # 打印表头
    print("\n" + "─" * (10 + sum(col_widths) + len(col_widths) * 3))
    header_line = "Excel行号 │ "
    for i, h in enumerate(header_display):
        header_line += f"{h:<{col_widths[i]}} │ "
    print(header_line)
    print("─" * len(header_line))
    
    # 打印数据行
    for idx, (row_num, row) in enumerate(display_rows[:MAX_DISPLAY_ROWS]):
        row_line = f"{row_num:<8} │ "
        for i in range(min(len(row), display_cols)):
            cell_str = str(row[i]) if i < len(row) and row[i] is not None else ""
            if len(cell_str) > col_widths[i]:
                cell_str = cell_str[:col_widths[i]-3] + "..."
            row_line += f"{cell_str:<{col_widths[i]}} │ "
        print(row_line)
    
    if len(display_rows) > MAX_DISPLAY_ROWS:
        print(f"\n... 还有 {len(display_rows) - MAX_DISPLAY_ROWS:,} 条结果未显示")
    
    print("\n" + "─" * (10 + sum(col_widths) + len(col_widths) * 3))
    
    # 显示统计信息
    if hidden_count > 0:
        print(f"\n💡 提示：原表格共有 {original_col_count} 列，自动隐藏了 {hidden_count} 个空白列")
        print(f"   如需显示全部列（包括空白列），请在代码中设置 HIDE_EMPTY_COLUMNS = False")
    
    return True

def save_results_filtered(file_path, sheet_name, headers, matched_rows, search_term, match_mode):
    """保存搜索结果到新文件（保存全部列，包括空白列）"""
    base_name = os.path.basename(file_path)
    base_name_without_ext = os.path.splitext(base_name)[0]
    clean_search = re.sub(r'[\\/*?:"<>|]', '_', search_term)
    mode_suffix = "模糊" if match_mode == "fuzzy" else "精确"
    output_filename = f"{base_name_without_ext}_{mode_suffix}匹配_{clean_search}.xlsx"
    output_path = os.path.join(os.path.dirname(file_path), output_filename)
    
    try:
        print(f"\n💾 正在保存结果（共 {len(matched_rows):,} 行，{len(headers) if headers else 0} 列）...")
        
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = sheet_name
        
        # 写入表头（全部列，包括空白列）
        if headers:
            ws_new.append(headers)
        else:
            if matched_rows:
                col_count = len(matched_rows[0][1])
                ws_new.append([f"列{i+1}" for i in range(col_count)])
        
        # 写入匹配的行（全部列，包括空白列）
        for i, (_, row) in enumerate(matched_rows, 1):
            ws_new.append(tuple(row) if row else ())
            
            if i % 1000 == 0:
                print(f"   已保存 {i:,}/{len(matched_rows):,} 行...")
                gc.collect()
        
        wb_new.save(output_path)
        print(f"\n✅ 结果已保存到：{output_path}")
        print(f"   文件大小：{format_file_size(os.path.getsize(output_path))}")
        return True
    except Exception as e:
        print(f"\n❌ 保存失败：{e}")
        return False

def show_help():
    """显示帮助信息"""
    print("\n" + "=" * 80)
    print("命令说明：")
    print("-" * 80)
    print("  /sheet     - 切换Sheet页")
    print("  /file      - 重新选择文件")
    print("  /showall   - 显示全部列（仅本次搜索）")
    print("  /mode      - 切换匹配模式（模糊/精确）")
    print("  /help      - 显示此帮助信息")
    print("  /q         - 退出程序")
    print("-" * 80)
    print(f"当前匹配模式：{'模糊匹配（包含即匹配）' if current_match_mode == 'fuzzy' else '精确匹配（完全相等）'}")
    print(f"空白列过滤：{'已启用（自动隐藏空白列）' if HIDE_EMPTY_COLUMNS else '已禁用'}")
    print("=" * 80)

def main():
    global MAX_DISPLAY_COLS, MAX_DISPLAY_ROWS, current_match_mode, HIDE_EMPTY_COLUMNS
    
    print("=" * 80)
    print("Excel 表格搜索工具 v3.6 - 智能隐藏空白列版".center(70))
    print("=" * 80)
    
    # 显示当前配置
    print(f"\n📋 当前配置：")
    print(f"   - 匹配模式：{'模糊匹配（包含即匹配）' if current_match_mode == 'fuzzy' else '精确匹配（完全相等）'}")
    print(f"   - 空白列处理：{'自动隐藏完全空白的列' if HIDE_EMPTY_COLUMNS else '显示所有列（包括空白列）'}")
    print(f"   - 控制台显示列数：{'全部有数据的列' if MAX_DISPLAY_COLS == 0 else f'前{MAX_DISPLAY_COLS}列'}")
    print(f"   - 控制台最大显示行数：{MAX_DISPLAY_ROWS}")
    print(f"   - 单元格最大显示宽度：{MAX_CELL_WIDTH}字符")
    print(f"   - 连续空白行终止阈值：{CONSECUTIVE_EMPTY_THRESHOLD}行")
    print(f"\n💡 提示：输入 '/help' 查看所有命令，输入 '/mode' 切换匹配模式")
    
    folder_path = get_folder_path()
    if folder_path is None:
        print("\n已退出程序。")
        return
    
    excel_files = get_excel_files(folder_path)
    
    while True:
        selected_file = select_excel_file(excel_files, folder_path)
        
        if selected_file == 'RESELECT_FOLDER':
            folder_path = get_folder_path()
            if folder_path is None:
                print("\n已退出程序。")
                return
            excel_files = get_excel_files(folder_path)
            continue
        elif selected_file is None:
            print("\n已退出程序。")
            return
        
        file_path = os.path.join(folder_path, selected_file)
        
        try:
            from openpyxl import load_workbook
            wb_temp = load_workbook(file_path, read_only=True)
            sheet_names = wb_temp.sheetnames
            wb_temp.close()
            
            print(f"\n✅ 文件：{selected_file}")
            print(f"   包含 {len(sheet_names)} 个 Sheet 页")
            
            current_sheet = select_sheet_with_smart_filter(file_path, sheet_names)
            if current_sheet is None:
                print("\n返回文件选择...")
                continue
            
            print(f"\n✅ 已选择 Sheet：{current_sheet}")
            
            while True:
                print(f"\n{'='*80}")
                print(f"当前模式：{'🔍 模糊匹配（包含即匹配）' if current_match_mode == 'fuzzy' else '🎯 精确匹配（完全相等）'}")
                print(f"空白列：{'✨ 自动隐藏' if HIDE_EMPTY_COLUMNS else '📊 显示全部'}")
                print(f"💡 提示：输入 '/mode' 切换匹配模式，'/help' 查看所有命令")
                print("-" * 80)
                search_term = input("请输入要查找的内容：").strip()
                
                # 处理特殊命令
                if search_term.lower() == '/help':
                    show_help()
                    continue
                elif search_term.lower() == '/showall':
                    MAX_DISPLAY_COLS = 0
                    print("\n✅ 已切换为显示全部列模式")
                    continue
                elif search_term.lower() == '/mode':
                    # 切换匹配模式
                    if current_match_mode == "fuzzy":
                        current_match_mode = "exact"
                        print("\n✅ 已切换为【精确匹配】模式（需要完全相等）")
                    else:
                        current_match_mode = "fuzzy"
                        print("\n✅ 已切换为【模糊匹配】模式（包含即可）")
                    continue
                elif search_term.lower() == 'q':
                    print("\n已退出程序。")
                    return
                elif search_term.lower() == '/sheet':
                    current_sheet = select_sheet_with_smart_filter(file_path, sheet_names)
                    if current_sheet:
                        print(f"\n✅ 已切换到 Sheet：{current_sheet}")
                    continue
                elif search_term.lower() == '/file':
                    print("\n返回文件选择...")
                    break
                elif not search_term:
                    print("❌ 查找内容不能为空")
                    continue
                
                try:
                    headers, matched_rows, total_valid_rows = search_in_sheet_smart(
                        file_path, current_sheet, search_term, current_match_mode
                    )
                    
                    if display_results_with_config(headers, matched_rows, search_term, total_valid_rows, current_match_mode):
                        print(f"\n{'='*80}")
                        save_choice = input("是否将结果保存到新 Excel 文件？(y/n)：").strip().lower()
                        if save_choice == 'y' and matched_rows:
                            save_results_filtered(file_path, current_sheet, headers, matched_rows, search_term, current_match_mode)
                    
                    del matched_rows
                    gc.collect()
                    
                except MemoryError:
                    print("\n❌ 内存不足！请尝试使用更精确的搜索词。")
                    gc.collect()
                except Exception as e:
                    print(f"\n❌ 搜索过程中发生错误：{e}")
                    import traceback
                    traceback.print_exc()
                    
        except Exception as e:
            print(f"\n❌ 读取文件失败：{e}")
            continue

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n用户取消操作")
    except Exception as e:
        print(f"\n发生未预期的错误：{e}")
        import traceback
        traceback.print_exc()
        input("\n按回车键退出...")